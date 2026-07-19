import openpyxl

def dump_excel(filepath):
    print("--- Formulas ---")
    wb_f = openpyxl.load_workbook(filepath, data_only=False)
    ws_f = wb_f.active
    
    formulas = {}
    for row in ws_f.iter_rows():
        for cell in row:
            if cell.data_type == 'f' or str(cell.value).startswith('='):
                formulas[cell.coordinate] = cell.value

    print("--- Values ---")
    wb_v = openpyxl.load_workbook(filepath, data_only=True)
    ws_v = wb_v.active
    
    values = {}
    for row in ws_v.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                values[cell.coordinate] = cell.value

    # Let's print out the important named ranges or just a combined view of all cells
    print("Combined View (Cell: Value | Formula):")
    all_cells = set(formulas.keys()).union(set(values.keys()))
    
    for cell_coord in sorted(all_cells, key=lambda x: (x[0], int(x[1:])) if x[1:].isdigit() else (x, 0)):
        val = values.get(cell_coord, "")
        form = formulas.get(cell_coord, "")
        if form or type(val) in (int, float):
            print(f"{cell_coord}: {val} | {form}")

    print("\nNamed Ranges:")
    for name in wb_f.defined_names.definedName:
        # dest is like 'Sheet1'!$A$1:$B$2
        dest = name.value
        print(f"{name.name} -> {dest}")

if __name__ == "__main__":
    dump_excel("Nivra Goal - Compute SIP_or_StepUP_SIP v3.xlsm")
